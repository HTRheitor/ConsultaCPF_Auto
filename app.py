import openpyxl
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import time 

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

driver = webdriver.Chrome()
driver.get('https://site_exemplo.com')

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    
    time.sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    time.sleep(2)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    time.sleep(2)
    botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    time.sleep(2)
    botao_pesquisar.click()
    time.sleep(5)
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    
    if status.text == 'em dia': 
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])

        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        
        planilha_fechamento.save('planilha fechamento.xlsx')
